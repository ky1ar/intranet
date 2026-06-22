import logging, os, uuid
import openpyxl
from docx import Document
from datetime import datetime, timedelta, date
from werkzeug.utils import secure_filename
from flask import request, send_file
from flask_jwt_extended import get_jwt_identity
from application.handlers import handle_exceptions
from application.utils import format_time, calculate_passed_days, format_date, size, file_extension, allowed_extension, upload_path, format_name, format_datetime
from application.services.module_service import ModuleService
from application.repository.import_repository import ImportRepository
from application.repository.user_repository import UserRepository
from application.repository.client_repository import ClientRepository
from application.repository.schedule_repository import ScheduleRepository
from application.services.push_service import PushSender
from application import socketio
from config import Paths


IMONTHS = {
    1: "Ene", 2: "Feb", 3: "Mar", 4: "Abr", 5: "May", 6: "Jun",
    7: "Jul", 8: "Ago", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dic",
}


class ImportService:
    def __init__(self):
        self.import_repository = ImportRepository()
        self.client_repository = ClientRepository()
        self.user_repository = UserRepository()
        self.schedule_repository = ScheduleRepository()
        self.module_service = ModuleService()
        self.push_service = PushSender()


    def _has_perm(self, user_id, perm_slug):
        """Verifica permiso del módulo imports"""
        result, code = self.module_service.check_permission(user_id, 'imports', perm_slug)
        if code != 200:
            return False
        return result.get('granted', False) if isinstance(result, dict) else False


    def _display_port_name(self, shipment):
        if getattr(shipment, "custom_port_name", None):
            return shipment.custom_port_name.strip()
        if shipment.port and shipment.port.name:
            return shipment.port.name.title()
        return "-"


    def _display_incoterm_code(self, line):
        if getattr(line, "custom_incoterm_name", None):
            return line.custom_incoterm_name.strip()
        if line.incoterm and line.incoterm.code:
            return line.incoterm.code
        return "-"


    def _format_notify_date_es(self, value):
        if not value:
            return "-"

        if isinstance(value, datetime):
            dt = value.date()
        elif isinstance(value, date):
            dt = value
        else:
            dt = datetime.strptime(str(value)[:10], "%Y-%m-%d").date()

        months = ["Ene.", "Feb.", "Mar.", "Abr.", "May.", "Jun.", "Jul.", "Ago.", "Sep.", "Oct.", "Nov.", "Dic."]
        base = f"{dt.day} de {months[dt.month - 1]} de {dt.year}"

        today = date.today()
        if dt == today:
            return f"Hoy, {base}"
        if dt == (today + timedelta(days=1)):
            return f"Mañana, {base}"
        return base


    def _validate_port_payload(self, data):
        port_id = data.get("port_id")
        custom_port_name = (data.get("custom_port_name") or "").strip()

        if not port_id and not custom_port_name:
            return None, None, "Selecciona un puerto o ingresa uno manual"

        if custom_port_name:
            return None, custom_port_name, None

        try:
            return int(port_id), None, None
        except Exception:
            return None, None, "Puerto inválido"


    def _normalize_lines_payload(self, lines):
        if not isinstance(lines, list) or not lines:
            return None, "Agrega al menos una línea"

        normalized = []
        for i, row in enumerate(lines, start=1):
            provider_id = row.get("provider_id")
            incoterm_id = row.get("incoterm_id")
            custom_incoterm_name = (row.get("custom_incoterm_name") or "").strip()
            advance_payment_percent = row.get("advance_payment_percent")
            balance_days = row.get("balance_days")

            if not provider_id:
                return None, f"Línea {i}: selecciona un proveedor"

            try:
                provider_id = int(provider_id)
            except Exception:
                return None, f"Línea {i}: proveedor inválido"

            if not incoterm_id and not custom_incoterm_name:
                return None, f"Línea {i}: selecciona un Incoterm o ingresa uno manual"

            parsed_incoterm_id = None
            if incoterm_id:
                try:
                    parsed_incoterm_id = int(incoterm_id)
                except Exception:
                    return None, f"Línea {i}: Incoterm inválido"

            # if advance_payment_percent is None or advance_payment_percent == "":
            #     return None, f"Línea {i}: ingresa adelanto"

            # try:
            #     advance_payment_percent = float(advance_payment_percent)
            # except Exception:
            #     return None, f"Línea {i}: adelanto inválido"

            # if advance_payment_percent < 0 or advance_payment_percent > 100:
            #     return None, f"Línea {i}: el adelanto debe estar entre 0 y 100"

            # if balance_days is None or balance_days == "":
            #     return None, f"Línea {i}: ingresa saldo"

            # try:
            #     balance_days = int(balance_days)
            # except Exception:
            #     return None, f"Línea {i}: saldo inválido"

            # if balance_days < 0:
            #     return None, f"Línea {i}: saldo inválido"

            normalized.append({
                "provider_id": provider_id,
                "incoterm_id": parsed_incoterm_id,
                "custom_incoterm_name": custom_incoterm_name or None,
                "advance_payment_percent": advance_payment_percent,
                "balance_days": balance_days,
                "position": i,
                "po_number": (row.get("po_number") or "").strip() or None,
            })

        return normalized, None
    

    def _unique_keep_order(self, values):
        result = []
        seen = set()
        for v in values:
            if not v:
                continue
            key = v.strip() if isinstance(v, str) else v
            if key in seen:
                continue
            seen.add(key)
            result.append(v)
        return result

    
    def _compact_label(self, items):
        items = self._unique_keep_order(items)
        if not items:
            return "-"
        if len(items) == 1:
            return items[0]
        return f"{items[0]} (+{len(items)-1})"


    def _sorted_lines(self, shipment):
        return sorted((shipment.lines or []), key=lambda x: (x.position or 999999, x.id or 0))


    def _shipment_summary(self, shipment):
        lines = self._sorted_lines(shipment)

        provider_names = [l.provider.name for l in lines if getattr(l, "provider", None)]
        incoterm_codes = [self._display_incoterm_code(l) for l in lines]

        provider_names = self._unique_keep_order(provider_names)
        incoterm_codes = self._unique_keep_order(incoterm_codes)

        return {
            "line_count": len(lines),
            "providers_text": self._compact_label(provider_names),
            "providers_full_text": ", ".join(provider_names) if provider_names else "-",
            "provider_names": provider_names,
            "incoterms_text": self._compact_label(incoterm_codes),
        }


    def _serialize_shipment_line(self, line):
        return {
            "id": line.id,
            "position": line.position,

            "provider_id": line.provider_id,
            "provider_name": line.provider.name if line.provider else None,
            "provider_contact": line.provider.contact if line.provider else None,
            "provider_email": line.provider.email if line.provider else None,
            "provider_phone": line.provider.phone if line.provider else None,
            "provider_telephone": line.provider.telephone if line.provider else None,

            "incoterm_id": line.incoterm_id,
            "custom_incoterm_name": getattr(line, "custom_incoterm_name", None),
            "incoterm_code": self._display_incoterm_code(line),

            "advance_payment_percent": float(line.advance_payment_percent or 0),
            "balance_days": line.balance_days or 0,
            "po_number": getattr(line, "po_number", None) or "",
        }


    @handle_exceptions
    def status(self):
        status, sc = self.import_repository.get_statuses()
        if sc != 200:
            return status, sc
        
        status_list = [
            {
                "id": option.id,
                "name": option.name,
                "image": option.image,

            } for option in status
        ]
        return status_list, 200
    
    
    @handle_exceptions
    def get_by_status(self, status_id):
        imports, isc = self.import_repository.get_by_status(status_id)
        if isc != 200:
            return imports, isc
        
        result = []
        for item in imports:
            summary = self._shipment_summary(item)

            order_data = {
                "id": item.id,
                "status_id": item.status_id,
                "fcl": int(getattr(item, "fcl", 0) or 0),

                # shared
                "business_name": item.business.name if item.business else "-",
                "port_name": item.port.name.title() if item.port and item.port.name else "-",
                "type_id": item.type_id,

                "local_agent_name": item.local_agent_name,
                "origin_agent_name": item.origin_agent_name,
                "advance_payment_percent": float(item.advance_payment_percent or 0),
                "balance_days": item.balance_days or 0,

                # lines summary
                "line_count": summary["line_count"],
                "providers_text": summary["providers_text"],
                "providers_full_text": summary["providers_full_text"],
                "provider_names": summary["provider_names"],
                "incoterms_text": summary["incoterms_text"],
            }

            result.append(order_data)

        return result, 200


    def _serialize_attachment(self, r):
        return {
            "id": r.id,
            "target": r.target,
            "original_name": r.original_name,
            "mime_type": r.mime_type,
            "size_bytes": r.size_bytes,
            "size_h": size(r.size_bytes),
            "created_at": str(r.created_at),
            "ext": file_extension(r.original_name),

            "inline_url": f"/imports/attachment/{r.id}?disposition=inline",
            "download_url": f"/imports/attachment/{r.id}?disposition=attachment",
            "preview_url": f"/imports/attachment/{r.id}/preview",
        }

    @handle_exceptions
    def dashboard(self):
        user_id = int(get_jwt_identity())
        user, uc = self.user_repository.get_user_by_id(user_id)
        if uc != 200:
            return user, uc

        imports, cc = self.import_repository.get_imports() 
        if cc != 200:
            return imports, cc

        statuses, sc = self.import_repository.get_statuses() 
        if sc != 200:
            return statuses, sc
        
        visible_statuses = [s for s in statuses if s.id not in (1, 14)]

        # Historial en UNA sola consulta: created_at por (import, estado actual) y
        # (import, estado inicial=1). Evita el N+1 (2 queries por import).
        all_ids = [imp.id for imp in imports]
        needed_status_ids = list({s.id for s in visible_statuses} | {1})
        hist_map, _ = self.import_repository.get_history_created_map(all_ids, needed_status_ids)
        today = date.today()

        result = []
        for i, status in enumerate(visible_statuses):
            is_last = (i == len(visible_statuses) - 1)

            import_status = [import_row for import_row in imports if import_row.status_id == status.id]

            imports_list = []
            for item in import_status:
                passed_days = 0
                updated_today = False

                current_created = hist_map.get((item.id, status.id))
                if current_created and current_created.date() == today:
                    updated_today = True

                first_created = hist_map.get((item.id, 1))
                if is_last:
                    if first_created and current_created:
                        passed_days = calculate_passed_days(first_created, current_created)
                else:
                    if first_created:
                        passed_days = calculate_passed_days(first_created)

                summary = self._shipment_summary(item)

                imports_data = {
                    "id": item.id,
                    "type_letter": "SP",
                    "updated_today": updated_today,
                    "passed_days": passed_days,
                    "status_id": status.id,
                    "traffic_light": item.traffic_light,
                    "fcl": int(getattr(item, "fcl", 0) or 0),

                    # shared
                    "business_name": item.business.name if item.business else "-",
                    "port_name": item.port.name.title() if item.port and item.port.name else "-",
                    "type_id": item.type_id,
                    # "advance_payment_percent": float(item.advance_payment_percent or 0),
                    # "balance_days": item.balance_days or 0,

                    # lines summary
                    "line_count": summary["line_count"],
                    "providers_text": summary["providers_text"],
                    "providers_full_text": summary["providers_full_text"],
                    "provider_names": summary["provider_names"],
                    "incoterms_text": summary["incoterms_text"],
                }
                    
                imports_list.append(imports_data)
            
            imports_list.sort(key=lambda x: x["passed_days"], reverse=True)
            if is_last:
                imports_list = imports_list[:10]  # tope solo en el último estado (acumula muchos)

            result.append({
                "status_id": status.id,
                "status_name": status.name,
                "imports": imports_list
            })

        return result, 200
    

    @handle_exceptions
    def options_provider(self):
        providers, cc = self.import_repository.get_options_provider() 
        if cc != 200:
            return providers, cc
        
        result = []
        for provider in providers:
            result.append({
                "id": provider.id,
                "name": provider.name
            })
        return result, 200


    @handle_exceptions
    def options_type(self):
        types, tc = self.import_repository.get_options_type() 
        if tc != 200:
            return types, tc
        
        result = []
        for item in types:
            result.append({
                "id": item.id,
                "name": item.name
            })
        return result, 200


    @handle_exceptions
    def options_business(self):
        business, cc = self.import_repository.get_options_business() 
        if cc != 200:
            return business, cc
        
        result = []
        for item in business:
            result.append({
                "id": item.id,
                "name": item.name
            })
        return result, 200
    

    @handle_exceptions
    def options_incoterm(self):
        incoterms, cc = self.import_repository.get_options_incoterm() 
        if cc != 200:
            return incoterms, cc
        
        result = []
        for incoterm in incoterms:
            result.append({
                "id": incoterm.id,
                "code": incoterm.code
            })
        return result, 200


    @handle_exceptions
    def options_port(self):
        ports, cc = self.import_repository.get_options_port() 
        if cc != 200:
            return ports, cc
        
        result = []
        for port in ports:
            result.append({
                "id": port.id,
                "name": port.name.title()
            })
        return result, 200
    

    @handle_exceptions
    def get_provider(self, provider_id):
        provider, pc = self.import_repository.get_provider(provider_id)
        if pc != 200:
            return provider, pc

        response = {
            "id": provider.id,
            "name": provider.name,
            "contact": provider.contact,
            "letter": provider.name[0].title(),
            "email": provider.email,
            "phone": provider.phone,
            "telephone": provider.telephone,
        }
        return response, 200
    

    def _format_list_item(self, item):
        summary = self._shipment_summary(item)
        reg, rc = self.import_repository.get_import_history(item.id, 1)
        created_at = format_datetime(reg.created_at) if rc == 200 and reg else None
        return {
            "id": item.id,
            "status_id": item.status_id,
            "status_name": item.status.name if item.status else None,
            "business_name": item.business.name if item.business else "-",
            "port_name": item.port.name.title() if item.port and item.port.name else "-",
            "providers_text": summary["providers_text"],
            "providers_full_text": summary["providers_full_text"],
            "line_count": summary["line_count"],
            "local_agent_name": item.local_agent_name.title() if item.local_agent_name else None,
            "origin_agent_name": item.origin_agent_name.title() if item.origin_agent_name else None,
            "created_at": created_at,
        }

    @handle_exceptions
    def search_imports(self, term):
        term = (term or "").strip()
        if len(term) < 2:
            return [], 200
        rows, rc = self.import_repository.search_imports(term)
        if rc != 200:
            return rows, rc
        return [self._format_list_item(r) for r in rows], 200

    @handle_exceptions
    def history(self, data):
        page     = data.get("page", 1)
        per_page = data.get("per_page", 12)
        result, rc = self.import_repository.get_imports_paginated(page, per_page)
        if rc != 200:
            return result, rc
        return {
            "list": [self._format_list_item(r) for r in result["list"]],
            "pagination": {
                "total":    result["total"],
                "page":     result["page"],
                "per_page": result["per_page"],
                "pages":    result["pages"],
            },
        }, 200

    @handle_exceptions
    def statistics(self):
        total, _         = self.import_repository.stats_total()
        active, _        = self.import_repository.stats_active()
        arriving, _      = self.import_repository.stats_arriving()
        containers, _    = self.import_repository.stats_containers()
        eta_rows, _      = self.import_repository.stats_arrivals_by_month()
        provider_rows, _ = self.import_repository.stats_by_provider()
        port_rows, _     = self.import_repository.stats_by_port()

        by_month = [
            {"period": f"{IMONTHS.get(int(p.split('-')[1]), p)} {p[:4]}", "count": c}
            for p, c in eta_rows
        ]
        by_provider = [{"provider": name, "count": c} for name, c in provider_rows]
        by_port = [{"port": name, "count": c} for name, c in port_rows]

        return {
            "count": {
                "total": total,
                "active": active,
                "arriving": arriving,
                "containers": containers,
            },
            "by_month": by_month,
            "by_provider": by_provider,
            "by_port": by_port,
        }, 200

    @handle_exceptions
    def view(self, import_id):
        import_shipment, isc = self.import_repository.get_import(import_id) 
        if isc != 200:
            return import_shipment, isc

        register_days = 0
        import_history, ihc = self.import_repository.get_import_history(import_id, 1) 
        if ihc == 200:
            register_days = calculate_passed_days(import_history.created_at)

        sorted_lines = self._sorted_lines(import_shipment)
        summary = self._shipment_summary(import_shipment)

        order_data = {
            "id": import_shipment.id,
            "passed_days": f"{register_days}  día{'' if register_days == 1 else 's'}",
            "status_id": import_shipment.status_id,
            "status_name": import_shipment.status.name,
            "summary": summary,

            "business_id": import_shipment.business_id,
            "business_name": import_shipment.business.name if import_shipment.business else None,
            "type_id": import_shipment.type_id,
            "type_name": import_shipment.type.name if import_shipment.type else None,
            "fcl": int(getattr(import_shipment, "fcl", 0) or 0),
            
            "port_name": self._display_port_name(import_shipment),
            "port_id": import_shipment.port_id,
            "custom_port_name": getattr(import_shipment, "custom_port_name", None),
            "tracking_link": getattr(import_shipment, "tracking_link", None),

            "local_agent_name": import_shipment.local_agent_name.title() if import_shipment.local_agent_name else None,
            "origin_agent_name": import_shipment.origin_agent_name.title() if import_shipment.origin_agent_name else None,
            "reference": getattr(import_shipment, "reference", None) or "",

            "advance_payment_percent": float(import_shipment.advance_payment_percent or 0),
            "balance_days": import_shipment.balance_days or 0,

            "tax_amount": float(import_shipment.tax_amount) if import_shipment.tax_amount is not None else None,
            "tax_currency": getattr(import_shipment, "tax_currency", None) or "PEN",
            "perception_amount": float(import_shipment.perception_amount) if import_shipment.perception_amount is not None else None,
            "perception_currency": getattr(import_shipment, "perception_currency", None) or "PEN",
            "dua_number": getattr(import_shipment, "dua_number", None) or "",

            "lines": [self._serialize_shipment_line(line) for line in sorted_lines],

            "dates": {
                "booking": format_date(import_shipment.booking_date),
                "etd": format_date(import_shipment.etd_date),
                "eta": format_date(import_shipment.eta_date),
                "deadline": format_date(import_shipment.deadline_date),
                "pay": format_date(import_shipment.pay_date),
            },
            "cargo_details": {
                "qty": import_shipment.qty,
                "pallets": import_shipment.pallets,
                "weight": import_shipment.weight,
                "volume": import_shipment.volume,
            },
            "traffic_light": import_shipment.traffic_light,
            "delivery": {
                "date": format_date(import_shipment.delivery_date),
                "time": format_time(import_shipment.delivery_time),
                "name": import_shipment.delivery_name,
                "phone": import_shipment.delivery_phone,
                "code": import_shipment.delivery_code,
            },
            "chats": []
        }

        for chat in import_shipment.chats:
            order_data["chats"].append({
                "id": chat.id,
                "comment": chat.comment,
                "commenter_id": chat.commenter_id,
                "commenter_name": format_name(chat.commenter.name),
                "commenter_image": chat.commenter.image,
                "created_at": format_datetime(chat.created_at),
            })

        attachments, arc = self.import_repository.get_attachments_by_import(import_id)
        if arc != 200:
            return attachments, arc

        target_labels = {
            "consolidated": "Documentos consolidados",
            "product_list": "Lista de productos",
            "invoice": "Invoice",
            "packing_list": "Packing List",
            "certificate": "Constancia",
            "coo": "COO",
            "arrive": "Documentos de llegada",
            "invoices": "Facturas",
            "default": "Archivos",
        }

        grouped = {}
        for r in attachments:
            key = (r.target or "default").strip() or "default"
            grouped.setdefault(key, []).append(self._serialize_attachment(r))

        ordered_targets = [
            "consolidated", "product_list", "invoice", "packing_list",
            "certificate", "coo", "arrive", "invoices", "default",
        ]

        attachment_sections = []
        for target in ordered_targets:
            files = grouped.get(target, [])
            if files:
                attachment_sections.append({
                    "target": target,
                    "label": target_labels.get(target, target.replace("_", " ").title()),
                    "files": files
                })

        for target, files in grouped.items():
            if target not in ordered_targets and files:
                attachment_sections.append({
                    "target": target,
                    "label": target_labels.get(target, target.replace("_", " ").title()),
                    "files": files
                })

        order_data["attachment_sections"] = attachment_sections
        return order_data, 200
    

    @handle_exceptions
    def new(self, data):
        user_id = int(get_jwt_identity())
        if not self._has_perm(user_id, 'add'):
            return "No autorizado", 403
    
        business_id = data.get("business_id")
        type_id = data.get("type_id")
        local_agent = (data.get("local_agent") or "").strip()
        origin_agent = (data.get("origin_agent") or "").strip()
        reference = (data.get("reference") or "").strip() or None
        lines = data.get("lines")
        fcl = 1 if str(data.get("fcl", 0)) == "1" else 0

        if not business_id:
            return "Selecciona una empresa", 400
        if not type_id:
            return "Selecciona un tipo de envío", 400
        if not local_agent:
            return "Ingresa un agente local", 400
        if not origin_agent:
            return "Ingresa un agente de origen", 400

        port_id, custom_port_name, port_error = self._validate_port_payload(data)
        if port_error:
            return port_error, 400

        normalized_lines, lines_error = self._normalize_lines_payload(lines)
        if lines_error:
            return lines_error, 400

        data["port_id"] = port_id
        data["custom_port_name"] = custom_port_name
        data["fcl"] = fcl
        data["reference"] = reference
        data["lines"] = normalized_lines

        import_id, ncc = self.import_repository.new_import(data)
        if ncc != 200:
            return import_id, ncc

        new_history, nhc = self.import_repository.new_history(import_id, user_id, 0, "Agregado a borrador")
        if nhc != 200:
            return new_history, nhc

        socketio.emit("imports_dashboard_update", {})
        return "Registrado correctamente", 200
    

    @handle_exceptions
    def move(self, data):
        user_id = int(get_jwt_identity())
        if not self._has_perm(user_id, 'move'):
            return "No autorizado", 403
        import_id = data.get("import_id")
        notes = data.get("notes")
        etd_date = data.get("etd_date")
        eta_date = data.get("eta_date")
        
        import_shipment, isc = self.import_repository.get_import(import_id) 
        if isc != 200:
            return import_shipment, isc

        summary = self._shipment_summary(import_shipment)

        current_status_id = import_shipment.status_id
        move, mc = self.import_repository.move_status(import_shipment, current_status_id, data) 
        if mc != 200:
            return move, mc
        
        new_history, nhc = self.import_repository.new_history(import_id, user_id, current_status_id, notes) 
        if nhc != 200:
            return new_history, nhc

        if current_status_id == 5:
            eta_base = str(eta_date)[:10]
            eta_dt = datetime.strptime(eta_base, "%Y-%m-%d")

            providers = ", ".join(summary["provider_names"]) if summary["provider_names"] else "Importación"

            payload = {
                "user_id": user_id,
                "title": f"{providers} | SP-{import_id}",
                "description": "Evento creado automáticamente para coordinación interna.",
                "start_datetime": eta_dt.strftime("%Y-%m-%dT00:00"),
                "end_datetime": eta_dt.strftime("%Y-%m-%dT00:00"),
                "meet": None,
                "hex_color": "#7986CB",
                "visibility_id": "1",
                "repeat_id": 1,
                "notify_id": 1,
                "all_day": True,
                "import_shipment_id": import_id,
            }
            
            new_event_id, aec = self.schedule_repository.add_event(payload)
            if aec != 200:
                return new_event_id, aec
            
            etd_text = self._format_notify_date_es(etd_date)
            eta_text = self._format_notify_date_es(eta_date)

            eta_prefix = "" if eta_text.startswith(("Hoy", "Mañana")) else "el "

            participants, _ = self.user_repository.get_all_user_ids()
            title = f"🛳️ Próxima salida"
            body = (
                f"{providers} SP-{import_id} tiene salida estimada {etd_text}, "
                f"y llegada a nuestro almacén {eta_prefix}{eta_text}. "
                f"Entra al calendario para ver el detalle de la importación."
            )

            registration_tokens, _ = self.push_service.prefetch_registration_tokens(participants)
            socketio.start_background_task(self.push_service.send_to_tokens, registration_tokens, title, body, None)
            socketio.start_background_task(socketio.emit, "calendar_update_dashboard", {})

        socketio.start_background_task(socketio.emit, "imports_dashboard_update", {})

        return "Actualizado correctamente", 200


    @handle_exceptions
    def down(self, data):
        user_id = int(get_jwt_identity())
        if not self._has_perm(user_id, 'move'):
            return "No autorizado", 403
        
        import_id = data.get("import_id")
        
        import_shipment, isc = self.import_repository.get_import(import_id) 
        if isc != 200:
            return import_shipment, isc

        current_status_id = import_shipment.status_id
        move, mc = self.import_repository.move_status(import_shipment, current_status_id, data, down=True) 
        if mc != 200:
            return move, mc
        
        new_history, nhc = self.import_repository.new_history(import_id, user_id, current_status_id, down=True) 
        if nhc != 200:
            return new_history, nhc

        socketio.emit("imports_dashboard_update", {})
        return "Actualizado correctamente", 200
    
    
    def attachments_list(self, import_id):
        target = (request.args.get("target") or "default").strip()

        attachments, rc = self.import_repository.get_attachments(import_id, target)
        if rc != 200:
            return attachments, rc

        data = []
        for r in attachments:
            data.append({
                "id": r.id,
                "target": r.target,
                "original_name": r.original_name,
                "mime_type": r.mime_type,
                "size_bytes": r.size_bytes,
                "size_h": size(r.size_bytes),
                "created_at": str(r.created_at),
                "ext": file_extension(r.original_name),

                "inline_url": f"/imports/attachment/{r.id}?disposition=inline",
                "download_url": f"/imports/attachment/{r.id}?disposition=attachment",
                "preview_url": f"/imports/attachment/{r.id}/preview",
            })

        return {"data": data}, 200


    def attachments_upload(self):
        user_id = int(get_jwt_identity())

        import_id = request.form.get("import_id")
        target = (request.form.get("target") or "default").strip()
        files = request.files.getlist("files[]")
        if not files:
            return {"data": {"message": "No files received"}}, 400

        max_bytes = Paths.MAX_BYTES
        saved_ids = []

        for f in files:
            if not f or not f.filename:
                continue

            if not allowed_extension(f.filename):
                return {"data": {"message": f"File type not allowed: {f.filename}"}}, 400

            size = None
            try:
                pos = f.stream.tell()
                f.stream.seek(0, os.SEEK_END)
                size = f.stream.tell()
                f.stream.seek(pos)
            except Exception:
                pass

            if size is not None and size > max_bytes:
                return {"data": {"message": f"File too large (max {max_bytes} bytes): {f.filename}"}}, 400

            original_name = f.filename
            ext = file_extension(original_name)
            safe = secure_filename(original_name) or f"file.{ext}"
            stored_name = f"{uuid.uuid4().hex}_{safe}"

            path = os.path.join(upload_path(Paths.IMPORTS), stored_name)
            f.save(path)

            mime = getattr(f, "mimetype", None)
            size_bytes = os.path.getsize(path)

            new_id, nc = self.import_repository.add_attachment(
                import_id=import_id,
                user_id=user_id,
                target=target,
                original_name=original_name,
                stored_name=stored_name,
                mime_type=mime,
                size_bytes=size_bytes,
            )
            if nc != 200:
                return new_id, nc

            saved_ids.append(new_id)

        return {"data": {"uploaded": saved_ids}}, 200


    def attachment_stream(self, attachment_id, disposition="inline"):
        row, rc = self.import_repository.get_attachment_by_id(int(attachment_id))
        if rc != 200:
            return row, rc

        path = os.path.join(Paths.IMPORTS, row.stored_name)
        if not os.path.exists(path):
            return "File missing on disk", 404

        as_attachment = (disposition == "attachment")
        return send_file(
            path,
            mimetype=row.mime_type or "application/octet-stream",
            as_attachment=as_attachment,
            download_name=row.original_name
        )


    def _file_ext(self, filename):
        return (filename.rsplit(".", 1)[-1] if "." in filename else "").lower()
    

    # GET /complaint/attachment/<attachment_id>/preview
    def attachment_preview(self, attachment_id):
        row, rc = self.import_repository.get_attachment_by_id(int(attachment_id))
        if rc != 200:
            return row, rc

        ext = self._file_ext(row.original_name)
        inline_url = f"/imports/attachment/{row.id}?disposition=inline"

        if ext in {"png","jpg","jpeg","webp","gif","pdf"}:
            return {"kind": "url", "url": inline_url, "name": row.original_name}, 200

        path = os.path.join(Paths.IMPORTS, row.stored_name)
        if not os.path.exists(path):
            return "File missing on disk", 404

        if ext in {"txt","xml"}:
            with open(path, "r", encoding="utf-8", errors="replace") as fp:
                text = fp.read(200_000)
            return {"kind": "text", "text": text, "name": row.original_name}, 200

        if ext == "docx" and Document:
            doc = Document(path)
            parts = [p.text for p in doc.paragraphs if p.text]
            text = "\n".join(parts)[:200_000]
            return {"kind": "text", "text": text, "name": row.original_name}, 200

        if ext == "xlsx" and openpyxl:
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active
            max_rows = min(ws.max_row or 0, 50)
            max_cols = min(ws.max_column or 0, 20)

            rows_html = []
            for r in range(1, max_rows + 1):
                cols = []
                for c in range(1, max_cols + 1):
                    v = ws.cell(row=r, column=c).value
                    cols.append(f"<td>{'' if v is None else str(v)}</td>")
                rows_html.append("<tr>" + "".join(cols) + "</tr>")

            html = "<table>" + "".join(rows_html) + "</table>"
            return {"kind": "html", "html": html, "name": row.original_name}, 200

        return {
            "kind": "download",
            "name": row.original_name,
            "message": "Vista previa no disponible",
            "download_url": f"/imports/attachment/{row.id}?disposition=attachment"
        }, 200


    @handle_exceptions
    def chat(self, data):
        import_id = data.get("import_id")
        if not import_id:
            return "import_id requerido", 400

        comment = (data.get("comment") or "").strip()
        if not comment:
            return "Comentario vacío", 400

        current_user_id = int(get_jwt_identity())
        user, uc = self.user_repository.get_user_by_id(current_user_id)
        if uc != 200:
            return user, uc
        
        user_name = user.name

        import_data, idc = self.import_repository.get_import(import_id)
        if idc != 200:
            return import_data, idc

        chat, cc = self.import_repository.add_chat(import_id=import_id, user_id=current_user_id, comment=comment)
        if cc != 200:
            return chat, cc

        participants, _ = self.import_repository.get_chat_participants(import_id=import_id, exclude_user_id=current_user_id)
        
        title = f"Nuevo mensaje, importación SP-{import_id}"
        body = f"{format_name(user_name, True)}: {comment}"

        registration_tokens, users_without = self.push_service.prefetch_registration_tokens(participants)

        socketio.start_background_task(self.push_service.send_to_tokens, registration_tokens, title, body, None)
        # socketio.start_background_task(socketio.emit, "imports_dashboard_update", {})

        if users_without:
            logging.info(f"[FCM] users_without_tokens={users_without}")

        return {
            "id": chat.id,
            "comment": chat.comment,
            "commenter_id": chat.commenter_id,
            "commenter_name": format_name(chat.commenter.name),
            "commenter_image": chat.commenter.image,
            "created_at": format_datetime(chat.created_at),
        }, 200


    @handle_exceptions
    def draft_update_agents(self, data):
        user_id = int(get_jwt_identity())
        if not self._has_perm(user_id, 'edit'):
            return "No autorizado", 403
    
        import_id = data.get("import_id")

        local_agent = (data.get("local_agent") or "").strip()
        origin_agent = (data.get("origin_agent") or "").strip()

        if not import_id:
            return "import_id requerido", 400
        if not local_agent:
            return "Ingresa agente local", 400
        if not origin_agent:
            return "Ingresa agente de origen", 400

        user, uc = self.user_repository.get_user_by_id(user_id)
        if uc != 200:
            return user, uc
        
        import_shipment, isc = self.import_repository.get_import(import_id)
        if isc != 200:
            return import_shipment, isc

        if import_shipment.status_id != 1:
            return "Solo se puede editar agentes cuando está en borrador", 400

        ok, rc = self.import_repository.update_draft_agents(
            import_shipment=import_shipment,
            local_agent=local_agent,
            origin_agent=origin_agent
        )
        if rc != 200:
            return ok, rc

        socketio.emit("imports_dashboard_update", {})
        return "Borrador actualizado correctamente", 200


    @handle_exceptions
    def draft_delete(self, data):
        user_id = int(get_jwt_identity())
        if not self._has_perm(user_id, 'edit'):
            return "No autorizado", 403
    
        import_id = data.get("import_id")
        if not import_id:
            return "import_id requerido", 400

        user, uc = self.user_repository.get_user_by_id(user_id)
        if uc != 200:
            return user, uc
        
        import_shipment, isc = self.import_repository.get_import(import_id)
        if isc != 200:
            return import_shipment, isc

        if import_shipment.status_id != 1:
            return "Solo se puede eliminar cuando está en borrador", 400

        # Limpieza de archivos físicos (por si existieran)
        attachments, arc = self.import_repository.get_attachments_by_import(import_id)
        if arc == 200:
            for a in attachments:
                try:
                    path = os.path.join(Paths.IMPORTS, a.stored_name)
                    if os.path.exists(path):
                        os.remove(path)
                except Exception as e:
                    logging.warning(f"No se pudo eliminar archivo {a.stored_name}: {e}")

        deleted, dc = self.import_repository.delete_import_draft(import_id)
        if dc != 200:
            return deleted, dc

        socketio.emit("imports_dashboard_update", {})
        return "Borrador eliminado correctamente", 200


    @handle_exceptions
    def update_basic(self, data):
        user_id = int(get_jwt_identity())
        import_id = data.get("import_id")

        if not import_id:
            return "import_id requerido", 400

        user, uc = self.user_repository.get_user_by_id(user_id)
        if uc != 200:
            return user, uc

        import_shipment, isc = self.import_repository.get_import(import_id)
        if isc != 200:
            return import_shipment, isc

        if import_shipment.status_id > 5:
            return "Solo se puede actualizar hasta el estado 5", 400

        local_agent = (data.get("local_agent") or "").strip()
        origin_agent = (data.get("origin_agent") or "").strip()
        reference = (data.get("reference") or "").strip() or None
        business_id = data.get("business_id")
        type_id = data.get("type_id")
        fcl = 1 if str(data.get("fcl", 0)) == "1" else 0

        if not local_agent:
            return "Ingresa agente local", 400
        if not origin_agent:
            return "Ingresa agente de origen", 400

        port_id, custom_port_name, port_error = self._validate_port_payload(data)
        if port_error:
            return port_error, 400

        normalized_lines, lines_error = self._normalize_lines_payload(data.get("lines"))
        if lines_error:
            return lines_error, 400

        payload = {
            "port_id": port_id,
            "business_id": business_id,
            "type_id": type_id,
            "fcl": fcl,
            "custom_port_name": custom_port_name,
            "local_agent_name": local_agent,
            "origin_agent_name": origin_agent,
            "reference": reference,
            "lines": normalized_lines,
        }

        ok, rc = self.import_repository.update_basic(import_shipment, payload)
        if rc != 200:
            return ok, rc

        socketio.emit("imports_dashboard_update", {})
        return "Importación actualizada correctamente", 200


    @handle_exceptions
    def confirm(self, data):
        user_id = int(get_jwt_identity())
        if not self._has_perm(user_id, 'edit'):
            return "No autorizado", 403
    
        import_id = data.get("import_id")
        if not import_id:
            return "import_id requerido", 400

        user, uc = self.user_repository.get_user_by_id(user_id)
        if uc != 200:
            return user, uc

        import_shipment, isc = self.import_repository.get_import(import_id)
        if isc != 200:
            return import_shipment, isc

        if import_shipment.status_id != 1:
            return "Solo se puede confirmar desde borrador", 400

        if import_shipment.type_id == 1:
            target_status_id = 2
        elif import_shipment.type_id == 2:
            target_status_id = 5
        else:
            return "Tipo de envío no soportado para confirmación", 400

        ok, rc = self.import_repository.set_status(import_shipment, target_status_id)
        if rc != 200:
            return ok, rc

        history, hc = self.import_repository.new_history_exact(
            import_id=import_id,
            user_id=user_id,
            status_id=target_status_id,
            notes="Importación confirmada"
        )
        if hc != 200:
            return history, hc

        socketio.emit("imports_dashboard_update", {})
        return "Importación confirmada correctamente", 200